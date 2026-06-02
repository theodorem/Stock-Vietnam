"""
VN Stock MA200 Screener
Dùng VPS API để lấy giá — không rate limit, không cần API key.
Tích lũy lịch sử giá trong data/ma200_prices.csv, tính MA200 mỗi ngày sau phiên.
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

import requests
import pandas as pd
from datetime import date
from pathlib import Path

# ── Cấu hình ─────────────────────────────────────────────────────────────────
DATA_DIR     = Path(__file__).parent / 'data'
PRICES_CSV   = DATA_DIR / 'ma200_prices.csv'
MA_SESSIONS  = 200
VOL_SESSIONS = 63
CHUNK_SIZE   = 400  # VPS API nhận tối đa 400 mã mỗi request

EXCHANGES = ['hose', 'hnx', 'upcom']
EXCHANGE_LABEL = {'hose': 'HOSE', 'hnx': 'HNX', 'upcom': 'UPCOM'}

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 '
                  '(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36',
    'Origin':  'https://banggia.vps.com.vn',
    'Referer': 'https://banggia.vps.com.vn/',
    'Accept':  'application/json, text/plain, */*',
}

# CORS proxies — GitHub Actions IP đôi khi bị VPS block, thử lần lượt
PROXY_FNS = [
    lambda u: u,
    lambda u: f'https://api.codetabs.com/v1/proxy?quest={u}',
    lambda u: f'https://api.allorigins.win/raw?url={u}',
]


# ── VPS API ──────────────────────────────────────────────────────────────────
def _get_json(url: str, timeout: int = 15):
    for proxy in PROXY_FNS:
        try:
            r = requests.get(proxy(url), headers=HEADERS, timeout=timeout)
            r.raise_for_status()
            return r.json()
        except Exception:
            continue
    raise RuntimeError(f'Không GET được: {url}')


def fetch_tickers() -> dict[str, str]:
    """Trả về {ticker: exchange_label} cho tất cả mã HOSE/HNX/UPCOM."""
    result = {}
    for ex in EXCHANGES:
        url = f'https://bgapidatafeed.vps.com.vn/getlistckindex/{ex}'
        try:
            data = _get_json(url)
            for item in data:
                # API trả về list string ["AAA","ABT",...] hoặc list dict [{"sym":"AAA",...}]
                if isinstance(item, str):
                    sym = item.strip()
                elif isinstance(item, dict):
                    sym = (item.get('sym') or item.get('ticker') or '').strip()
                else:
                    continue
                if sym:
                    result[sym] = EXCHANGE_LABEL[ex]
        except Exception as e:
            print(f'  Cảnh báo: không lấy được danh sách {ex.upper()} — {e}')
    return result


def fetch_prices(symbols: list[str]) -> pd.DataFrame:
    """Lấy giá đóng cửa + khối lượng của tất cả mã (chunk 400)."""
    rows = []
    for i in range(0, len(symbols), CHUNK_SIZE):
        chunk = symbols[i : i + CHUNK_SIZE]
        url = f'https://bgapidatafeed.vps.com.vn/getliststockdata/{",".join(chunk)}'
        try:
            data = _get_json(url)
            for item in data:
                close  = float(item.get('lastPrice', 0) or 0)
                volume = float(item.get('lot', 0) or 0) * 10  # VPS lot × 10 = cổ phiếu
                if close > 0:
                    rows.append({'ticker': item.get('sym', ''), 'close': close, 'volume': volume})
        except Exception as e:
            print(f'  Cảnh báo: lỗi chunk {i // CHUNK_SIZE + 1} — {e}')
    return pd.DataFrame(rows)


# ── Lịch sử giá ─────────────────────────────────────────────────────────────
def load_history() -> pd.DataFrame:
    if PRICES_CSV.exists():
        return pd.read_csv(PRICES_CSV, parse_dates=['date'])
    return pd.DataFrame(columns=['date', 'ticker', 'exchange', 'close', 'volume'])


def update_history(history: pd.DataFrame, today_df: pd.DataFrame,
                   ticker_exchange: dict, trading_date: date) -> pd.DataFrame:
    if today_df.empty:
        return history
    today_df = today_df.copy()
    today_df['date']     = pd.Timestamp(trading_date)
    today_df['exchange'] = today_df['ticker'].map(ticker_exchange).fillna('UNKNOWN')

    # Xóa dữ liệu cũ của ngày hôm nay (tránh duplicate khi chạy lại)
    history = history[history['date'] != pd.Timestamp(trading_date)]
    combined = pd.concat([history, today_df[['date','ticker','exchange','close','volume']]],
                         ignore_index=True)
    return combined.sort_values(['ticker', 'date'])


def save_history(df: pd.DataFrame):
    DATA_DIR.mkdir(exist_ok=True)
    df.to_csv(PRICES_CSV, index=False)


# ── Tính MA200 và lọc ────────────────────────────────────────────────────────
def run_screener(history: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for ticker, grp in history.groupby('ticker'):
        grp = grp.sort_values('date')
        closes  = grp['close'].to_numpy()
        volumes = grp['volume'].to_numpy()

        if len(closes) < MA_SESSIONS:
            continue

        last_close = float(closes[-1])
        if last_close <= 0:
            continue

        ma200 = float(closes[-MA_SESSIONS:].mean())
        if ma200 <= 0:
            continue

        # Điều kiện lọc: giá đóng cửa > MA200
        if last_close <= ma200:
            continue

        pct     = round((last_close / ma200 - 1) * 100, 2)
        vol_win = volumes[-VOL_SESSIONS:] if len(volumes) >= VOL_SESSIONS else volumes
        avg_vol = int(round(float(vol_win.mean()), 0)) if len(vol_win) > 0 else 0
        exchange = grp['exchange'].iloc[-1]

        rows.append({
            'Mã':              ticker,
            'Sàn':             exchange,
            'Giá đóng cửa':   last_close,
            'MA 200':          round(ma200, 2),
            '% trên MA200':   pct,
            'Khối lượng TB 3T': avg_vol,
        })

    if not rows:
        return pd.DataFrame()

    return pd.DataFrame(rows).sort_values('% trên MA200', ascending=False, ignore_index=True)


# ── Xuất Excel ───────────────────────────────────────────────────────────────
def export_excel(df: pd.DataFrame, trading_date: date) -> Path:
    from openpyxl.styles import Alignment, Font, PatternFill

    DATA_DIR.mkdir(exist_ok=True)
    path = DATA_DIR / f'ma200_results_{trading_date}.xlsx'

    with pd.ExcelWriter(path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='MA200 Screener')
        ws = writer.sheets['MA200 Screener']
        fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
        hdr_font = Font(bold=True, color='FFFFFF', name='Calibri')
        for cell in ws[1]:
            cell.fill = fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal='center')
        for col in ws.columns:
            w = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = max(w + 4, 14)
        ws.row_dimensions[1].height = 20

    return path


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    today = date.today()
    print('═' * 55)
    print('  VN Stock MA200 Screener  (VPS API)')
    print(f'  Ngày: {today}')
    print('═' * 55)

    # 1. Danh sách mã
    print('\n[1/4] Lấy danh sách cổ phiếu...')
    ticker_exchange = fetch_tickers()
    symbols = list(ticker_exchange.keys())
    print(f'     → {len(symbols)} mã (HOSE + HNX + UPCOM)')

    # 2. Giá đóng cửa hôm nay
    print('\n[2/4] Lấy giá đóng cửa từ VPS API...')
    today_df = fetch_prices(symbols)
    print(f'     → {len(today_df)} mã có giá')

    # 3. Cập nhật lịch sử
    print('\n[3/4] Cập nhật lịch sử...')
    history = load_history()
    history = update_history(history, today_df, ticker_exchange, today)
    save_history(history)
    days = history['date'].nunique()
    print(f'     → {days} phiên tích lũy / cần {MA_SESSIONS} để tính MA200')

    # 4. Tính MA200 và xuất
    print('\n[4/4] Tính MA200 và lọc...')
    results = run_screener(history)

    if results.empty:
        if days < MA_SESSIONS:
            print(f'  Cần thêm {MA_SESSIONS - days} phiên nữa để có đủ dữ liệu MA200.')
        else:
            print('  Không có mã nào đạt điều kiện (giá > MA200).')
        return

    path = export_excel(results, today)

    print('\n' + '═' * 55)
    print('  HOÀN THÀNH')
    print('═' * 55)
    print(f'  Mã đạt điều kiện : {len(results)}')
    print(f'  Phiên lịch sử    : {days}')
    print(f'  Kết quả          : data/{path.name}')
    print('═' * 55)


if __name__ == '__main__':
    main()
